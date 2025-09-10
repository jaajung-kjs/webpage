"""
Excel Transformer - 휴전계획 보고서 변환 모듈
Vercel Functions용으로 수정된 버전
"""

import pandas as pd
import numpy as np
from datetime import datetime
from html.parser import HTMLParser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')


class HTMLTableParser(HTMLParser):
    """HTML 테이블을 파싱하는 클래스"""
    
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_row = False
        self.in_cell = False
        self.rows = []
        self.current_row = []
        self.current_cell = ''
        
    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self.in_table = True
        elif tag == 'tr' and self.in_table:
            self.in_row = True
            self.current_row = []
        elif tag in ['td', 'th'] and self.in_row:
            self.in_cell = True
            self.current_cell = ''
            
    def handle_endtag(self, tag):
        if tag == 'table':
            self.in_table = False
        elif tag == 'tr' and self.in_row:
            self.in_row = False
            if self.current_row:
                self.rows.append(self.current_row)
        elif tag in ['td', 'th'] and self.in_cell:
            self.in_cell = False
            self.current_row.append(self.current_cell.strip())
            
    def handle_data(self, data):
        if self.in_cell:
            self.current_cell += data


def parse_html_buffer(buffer):
    """BytesIO 버퍼에서 HTML을 파싱하여 데이터프레임으로 변환"""
    
    # 버퍼 읽기 (EUC-KR 인코딩 시도)
    try:
        buffer.seek(0)
        content = buffer.read().decode('euc-kr')
    except UnicodeDecodeError:
        # EUC-KR이 실패하면 UTF-8 시도
        try:
            buffer.seek(0)
            content = buffer.read().decode('utf-8')
        except UnicodeDecodeError:
            # 그래도 실패하면 cp949 시도
            buffer.seek(0)
            content = buffer.read().decode('cp949', errors='ignore')
    
    # HTML 파싱
    parser = HTMLTableParser()
    parser.feed(content)
    
    if len(parser.rows) < 3:
        return pd.DataFrame()  # 빈 데이터프레임 반환
    
    # 데이터 구조화
    data_rows = []
    for row in parser.rows[3:]:  # 헤더 제외
        if len(row) >= 15:
            data_rows.append({
                '사업소_1차': row[0],
                '사업소_2차': row[1],
                '변전소': row[2],
                '전압': row[3],
                '설비명': row[4],
                '공사명': row[5],
                '공사개요': row[6],
                '휴전일시_시작': row[7],
                '휴전일시_종료': row[8],
                '구분': row[9],
                '주관부서': row[10],
                '감독자': row[11],
                '도급업체명': row[12] if len(row) > 12 else '',
                '수속절차': row[13] if len(row) > 13 else '',
                '휴전종류': row[14] if len(row) > 14 else ''
            })
    
    df = pd.DataFrame(data_rows)
    return df


def transform_data(df):
    """데이터를 출력 형식에 맞게 변환"""
    
    if df.empty:
        return pd.DataFrame()
    
    # 새로운 데이터프레임 생성
    output_data = []
    
    # 허용된 2차 값 정의
    allowed_second_values = ['직할', '강릉', '동해', '원주', '태백']
    
    for idx, row in df.iterrows():
        # 2차 값이 허용된 값인지 확인
        second_value = row['사업소_2차']
        if second_value not in allowed_second_values:
            continue  # 허용되지 않은 2차 값이면 건너뛰기
        
        # 수속절차가 "활선작업"인지 확인
        is_live_work = row.get('수속절차', '') == '활선작업'
        
        output_data.append({
            '구분': '활선' if is_live_work else '휴전',
            '순번': 0,  # 정렬 후 재할당
            '휴전일시_시작': row['휴전일시_시작'],
            '휴전일시_종료': row['휴전일시_종료'],
            '2차': second_value,
            '변전소': row['변전소'],
            '전압': row['전압'],
            '설비명': row['설비명'],
            '공사개요': row['공사개요'],
            '구분2': row['구분'],  # 휴전 구분 (연속, 당일 등)
            '주관부서': row['주관부서'],
            '감독자': row['감독자'],
            '안전관리자': ''  # 빈 컬럼
        })
    
    if not output_data:
        return pd.DataFrame()
    
    df_output = pd.DataFrame(output_data)
    
    # 2차 컬럼의 정렬 순서 정의
    second_order = {'직할': 1, '강릉': 2, '동해': 3, '원주': 4, '태백': 5}
    
    # 정렬을 위한 임시 컬럼 추가
    df_output['2차_순서'] = df_output['2차'].map(lambda x: second_order.get(x, 999))
    
    # 시작 시간을 datetime으로 변환
    df_output['시작_datetime'] = pd.to_datetime(df_output['휴전일시_시작'], errors='coerce')
    
    # 구분별로 그룹화하고 정렬
    df_sorted_list = []
    
    # 휴전 그룹 먼저 처리
    df_shutdown = df_output[df_output['구분'] == '휴전'].copy()
    if not df_shutdown.empty:
        df_shutdown = df_shutdown.sort_values(by=['2차_순서', '시작_datetime'])
        df_sorted_list.append(df_shutdown)
    
    # 활선 그룹 처리
    df_live = df_output[df_output['구분'] == '활선'].copy()
    if not df_live.empty:
        df_live = df_live.sort_values(by=['2차_순서', '시작_datetime'])
        df_sorted_list.append(df_live)
    
    # 정렬된 데이터 합치기
    if df_sorted_list:
        df_final = pd.concat(df_sorted_list, ignore_index=True)
    else:
        df_final = df_output
    
    # 임시 컬럼 제거
    df_final = df_final.drop(['2차_순서', '시작_datetime'], axis=1)
    
    # 순번 재할당
    df_final['순번'] = range(1, len(df_final) + 1)
    
    return df_final


def save_to_excel_buffer(df, report_date):
    """데이터프레임을 Excel BytesIO 버퍼로 저장 (서식 포함)"""
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    
    # 제목 추가
    title = f"일일 휴전계획 보고({report_date.strftime('%m.%d')})"
    ws.merge_cells('A1:M1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더 스타일 정의
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 활선 작업 행의 배경색 정의 (연한 노란색)
    live_work_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 첫 번째 헤더 행 (병합된 셀들)
    ws.merge_cells('A4:A5')
    ws['A4'] = '구분'
    ws.merge_cells('B4:B5')
    ws['B4'] = '순번'
    ws.merge_cells('C4:F4')
    ws['C4'] = '휴전일시'
    ws['C5'] = '시작'
    ws['D5'] = '종료'
    ws['E5'] = '2차'
    ws['F5'] = '변전소'
    ws.merge_cells('G4:G5')
    ws['G4'] = '전압'
    ws.merge_cells('H4:H5')
    ws['H4'] = '설비명'
    ws.merge_cells('I4:I5')
    ws['I4'] = '공사개요'
    ws.merge_cells('J4:J5')
    ws['J4'] = '구분'
    ws.merge_cells('K4:K5')
    ws['K4'] = '주관부서'
    ws.merge_cells('L4:L5')
    ws['L4'] = '감독자'
    ws.merge_cells('M4:M5')
    ws['M4'] = '안전관리자'
    
    # 헤더 스타일 적용
    for row in ws['A4:M5']:
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # 데이터 추가
    start_row = 6
    for idx, row in df.iterrows():
        current_row = start_row + idx
        
        # 데이터 입력
        ws.cell(row=current_row, column=1, value=row['구분'])
        
        # 순번 컬럼에 Excel 함수 입력
        ws.cell(row=current_row, column=2, value=f'=ROW()-5')
        
        ws.cell(row=current_row, column=3, value=row['휴전일시_시작'])
        ws.cell(row=current_row, column=4, value=row['휴전일시_종료'])
        ws.cell(row=current_row, column=5, value=row['2차'])
        ws.cell(row=current_row, column=6, value=row['변전소'])
        ws.cell(row=current_row, column=7, value=row['전압'])
        ws.cell(row=current_row, column=8, value=row['설비명'])
        ws.cell(row=current_row, column=9, value=row['공사개요'])
        ws.cell(row=current_row, column=10, value=row['구분2'])
        ws.cell(row=current_row, column=11, value=row['주관부서'])
        ws.cell(row=current_row, column=12, value=row['감독자'])
        ws.cell(row=current_row, column=13, value=row['안전관리자'])
        
        # 스타일 적용
        for col in range(1, 14):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            # 순번 컬럼(B열)은 중앙 정렬
            if col == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # 활선 작업인 경우 배경색 적용
            if row['구분'] == '활선':
                cell.fill = live_work_fill
    
    # 구분 컬럼 스타일 재적용
    for idx, row in df.iterrows():
        current_row = start_row + idx
        cell = ws[f'A{current_row}']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 컬럼 너비 조정
    column_widths = {
        'A': 8, 'B': 6, 'C': 18, 'D': 18, 'E': 10, 'F': 12,
        'G': 8, 'H': 25, 'I': 40, 'J': 8, 'K': 15, 'L': 20, 'M': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # BytesIO 버퍼에 저장
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer